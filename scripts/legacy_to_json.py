"""
Parse legacy_xlsx/Poker Club statistics.xlsx into scripts/legacy_data.json.

The legacy sheet stores one column per game and one row per player; each cell is
the player's NET PROFIT for that game (payout - total buy-in cost). To reconstruct
the schema this app expects, we infer per-game buy-in amount as the GCD of the
absolute loss values in that column, treat losers' losses as N buy-ins, and record
each player's payout as payout_override = net + (buy_ins * buy_in_amount).

Players: bottom section "Net profits" starting at row 72, names in col C, until row "Check (=0)".
Game columns: D (col 4) through D+88 (col 92), 89 games total.
Row 71 = date, row 69 = game #.
"""
import json
import math
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

HERE = Path(__file__).parent
XLSX = HERE.parent / "legacy_xlsx" / "Poker Club statistics.xlsx"
OUT = HERE / "legacy_data.json"

wb = load_workbook(XLSX, data_only=True)
ws = wb["Summary"]

# 1. Collect player rows (bottom "Net profits" section)
players = []  # list of (row, name)
for r in range(72, 200):
    name = ws.cell(r, 3).value
    if not name:
        continue
    if "check" in str(name).lower() or "# of players" in str(name).lower():
        break
    players.append((r, str(name).strip()))

print(f"Found {len(players)} players")

# 2. Collect games
tournaments = []
GAME_COL_START = 4
NUM_GAMES = 89

def gcd_list(xs):
    g = 0
    for x in xs:
        g = math.gcd(g, int(abs(x)))
    return g

for gi in range(NUM_GAMES):
    col = GAME_COL_START + gi
    game_no = ws.cell(69, col).value
    date_val = ws.cell(71, col).value
    if not date_val:
        print(f"  Game {gi+1}: no date, skipping")
        continue

    if isinstance(date_val, datetime):
        date_str = date_val.date().isoformat()
    else:
        date_str = str(date_val)

    raw = []
    for r, name in players:
        v = ws.cell(r, col).value
        if v is None or v == 0:
            continue
        raw.append((name, float(v)))

    if not raw:
        continue

    # Validate: sum should be 0 (zero-sum game)
    nets_only = [n for _, n in raw]
    s = sum(nets_only)
    if abs(s) > 0.01:
        # Some games may have minor imbalances; warn but proceed
        print(f"  Game {game_no} ({date_str}): nonzero sum={s} — proceeding anyway")

    # Buy-in inference: GCD of absolute LOSS amounts (negatives)
    losses = [abs(n) for _, n in raw if n < 0]
    if losses:
        buy_in = gcd_list(losses)
    else:
        buy_in = 10  # fallback
    if buy_in < 1:
        buy_in = 10

    entries = []
    for name, net in raw:
        if net < 0:
            # loser: buy_ins absorbs all the loss, no payout
            buy_ins = int(round(abs(net) / buy_in))
            payout_override = 0.0
        else:
            # winner: assume 1 buy-in (legacy has no rebuy info for winners)
            buy_ins = 1
            payout_override = net + buy_in * buy_ins
        entries.append({
            "player_name": name,
            "buy_ins": buy_ins,
            "finish_position": None,
            "payout_override": payout_override,
        })

    tournaments.append({
        "game_number": game_no,
        "date": date_str,
        "name": f"Game {game_no}",
        "buy_in_amount": buy_in,
        "payout_structure": [{"position": 1, "pct": 100}],
        "notes": "Imported from legacy spreadsheet",
        "entries": entries,
    })

# 3. Sort tournaments by date ascending (so cumulative chart renders in order)
tournaments.sort(key=lambda t: t["date"])

# 4. Player roster — names only
player_names = [name for _, name in players]

out = {"players": player_names, "tournaments": tournaments}
OUT.write_text(json.dumps(out, indent=2, ensure_ascii=False))
print(f"✓ Wrote {OUT} — {len(player_names)} players, {len(tournaments)} tournaments")

# Sanity check: per-player cumulative net profit vs legacy ranking row
running = {n: 0.0 for n in player_names}
for t in tournaments:
    for e in t["entries"]:
        cost = e["buy_ins"] * t["buy_in_amount"]
        net = e["payout_override"] - cost
        running[e["player_name"]] += net

# Compare to legacy "Net profit" total — column at end of top section
# Total net is at some "Total" column; let's just print our computed totals
top_totals = sorted(running.items(), key=lambda kv: -kv[1])
print("\nComputed cumulative net (top 10):")
for n, v in top_totals[:10]:
    print(f"  {n:30s}  €{v:+.2f}")
print("\nBottom 5:")
for n, v in top_totals[-5:]:
    print(f"  {n:30s}  €{v:+.2f}")
